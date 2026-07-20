from __future__ import annotations

import base64
import json
import re
import shutil
from pathlib import Path
from typing import Any

from croqui_engine.excel.native_graph_exporter import (
    NativeExcelTemplateUnavailable,
    export_native_graph_xlsx,
)
from croqui_engine.excel.official_template_assets import official_rge_logo_svg
from croqui_engine.graph.croqui_graph import CroquiGraph, validate_croqui_graph_for_export
from croqui_engine.office.libreoffice import convert_to_xls
from croqui_engine.symbols.official_symbol_assets import default_symbol_assets_dir


def export_from_croqui_graph_svg(
    graph: CroquiGraph,
    svg: str,
    output_dir: Path,
) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    graph.validation = validate_croqui_graph_for_export(graph, svg)
    graph_path = output_dir / "edited_croqui_graph.json"
    svg_path = output_dir / "croqui_final.svg"
    pdf_path = output_dir / "croqui_final.pdf"
    png_path = output_dir / "croqui_final.png"
    xls_path = output_dir / "croqui_final.xls"
    xlsx_path = output_dir / "croqui_final.xlsx"
    report_path = output_dir / "output_validation_report.json"

    svg = _embed_official_assets(svg)
    graph_path.write_text(
        json.dumps(graph.as_dict(), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    svg_path.write_text(svg, encoding="utf-8")
    _svg_to_pdf(svg, pdf_path)
    _pdf_to_png(pdf_path, png_path)
    try:
        export_native_graph_xlsx(graph, xlsx_path)
        converted_xls = convert_to_xls(xlsx_path, output_dir)
        if converted_xls != xls_path:
            shutil.copy2(converted_xls, xls_path)
    except (NativeExcelTemplateUnavailable, RuntimeError, OSError, ValueError) as exc:
        graph.validation.warnings.append(
            {
                "code": "NATIVE_EXCEL_FALLBACK",
                "message": str(exc),
            }
        )
        _write_xls_with_svg_image(graph, svg, xls_path)
        _write_xlsx_with_svg_image(graph, png_path, xlsx_path)
    _write_validation_report(graph, report_path)
    return {
        "graph": graph_path,
        "svg": svg_path,
        "pdf": pdf_path,
        "png": png_path,
        "xls": xls_path,
        "xlsx": xlsx_path,
        "validation_report": report_path,
    }


def _embed_official_assets(svg: str) -> str:
    assets_dir = default_symbol_assets_dir().resolve()

    def replace(match: re.Match[str]) -> str:
        quote = match.group("quote")
        filename = Path(match.group("path")).name
        asset = (assets_dir / filename).resolve()
        if asset.parent != assets_dir or not asset.is_file():
            return match.group(0)
        encoded = base64.b64encode(asset.read_bytes()).decode("ascii")
        return f'{match.group("attr")}={quote}data:image/png;base64,{encoded}{quote}'

    embedded = re.sub(
        r'(?P<attr>(?:xlink:)?href)=(?P<quote>["\'])(?P<path>/static/img/symbols/official/[^"\']+)(?P=quote)',
        replace,
        svg,
    )
    if "/api/assets/rge-logo.svg" in embedded:
        try:
            logo = base64.b64encode(official_rge_logo_svg().read_bytes()).decode("ascii")
            embedded = embedded.replace(
                "/api/assets/rge-logo.svg",
                f"data:image/svg+xml;base64,{logo}",
            )
        except (FileNotFoundError, RuntimeError, OSError):
            pass
    return embedded


def _svg_to_pdf(svg: str, output_path: Path) -> Path:
    import fitz

    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc = fitz.open("svg", svg.encode("utf-8"))
    try:
        output_path.write_bytes(doc.convert_to_pdf())
    finally:
        doc.close()
    return output_path


def _pdf_to_png(pdf_path: Path, output_path: Path) -> Path:
    import fitz

    with fitz.open(pdf_path) as doc:
        pix = doc[0].get_pixmap(matrix=fitz.Matrix(1.35, 1.35), alpha=False)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        pix.save(output_path)
    return output_path


def _svg_to_bmp(svg: str, output_path: Path) -> Path:
    import fitz
    from PIL import Image

    doc = fitz.open("svg", svg.encode("utf-8"))
    try:
        pdf_bytes = doc.convert_to_pdf()
    finally:
        doc.close()
    with fitz.open("pdf", pdf_bytes) as pdf:
        pix = pdf[0].get_pixmap(matrix=fitz.Matrix(1.4, 1.4), alpha=False)
        png_bytes = pix.tobytes("png")
    import io

    image = Image.open(io.BytesIO(png_bytes)).convert("RGB")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(output_path, format="BMP")
    return output_path


def _write_xls_with_svg_image(graph: CroquiGraph, svg: str, output_path: Path) -> Path:
    import xlwt

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Croqui")
    _configure_sheet(sheet)
    styles = _styles(xlwt)
    _write_header(sheet, graph, styles)
    for row in range(6, 30):
        sheet.write_merge(row, row, 1, 47, "", styles["drawing"])
    bmp = _svg_to_bmp(svg, output_path.parent / "croqui_final_from_editor.bmp")
    try:
        sheet.insert_bitmap(str(bmp), 7, 1, x=8, y=4, scale_x=0.62, scale_y=0.62)
    except Exception:
        sheet.write_merge(
            7, 29, 1, 47, "SVG final exportado pelo editor salvo ao lado deste XLS.", styles["note"]
        )
    _write_validation(sheet, graph, styles)
    workbook.save(str(output_path))
    return output_path


def _write_xlsx_with_svg_image(graph: CroquiGraph, png_path: Path, output_path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as WorksheetImage
    from openpyxl.styles import Alignment, Border, Font, Side

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Croqui"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    thin = Side(style="thin", color="222222")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 20):
        ws.column_dimensions[chr(64 + col)].width = 12
    for row in range(1, 38):
        ws.row_dimensions[row].height = 20
    ws.merge_cells("A1:C3")
    ws["A1"] = "RGE"
    ws["A1"].font = Font(name="Arial", bold=True, size=22, italic=True)
    headers = [
        ("A4", "C4", "Departamento:", graph.header.departamento),
        ("D4", "G4", "Município:", graph.header.municipio),
        ("H4", "L4", "Equipamento:", graph.header.equipamento),
        ("A5", "C5", "Data do Levantamento:", graph.header.data_levantamento),
        ("D5", "L5", "Responsável:", graph.header.responsavel),
    ]
    for start, end, label, value in headers:
        ws.merge_cells(f"{start}:{end}")
        cell = ws[start]
        cell.value = f"{label} {value}".strip()
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Arial", size=9, bold=label.endswith(":"))
    for row in range(6, 31):
        for col in range(1, 13):
            ws.cell(row=row, column=col).border = border
    image = WorksheetImage(str(png_path))
    image.width = 930
    image.height = 520
    ws.add_image(image, "A7")
    ws["A32"] = f"Status: {graph.validation.status}"
    ws["A33"] = "Avisos: " + "; ".join(
        str(item.get("code") or item)
        for item in [*graph.validation.blockingErrors, *graph.validation.warnings]
    )
    wb.save(output_path)
    return output_path


def _configure_sheet(sheet: Any) -> None:
    try:
        sheet.show_grid = False
        sheet.set_portrait(False)
        sheet.paper_size_code = 9
        sheet.set_fit_width_to_pages(1)
        sheet.set_fit_height_to_pages(1)
        sheet.set_header_str(b"")
        sheet.set_footer_str(b"")
    except Exception:
        pass
    for col in range(49):
        sheet.col(col).width = 720 if 1 <= col <= 47 else 300
    for row in range(46):
        sheet.row(row).height = 300


def _styles(xlwt: Any) -> dict[str, Any]:
    return {
        "label": _style(xlwt, bold=True, border=True),
        "value": _style(xlwt, border=True, align="center"),
        "drawing": _style(xlwt, border=True),
        "note": _style(xlwt, border=True, align="center", bold=True),
        "warning": _style(xlwt, border=True, wrap=True),
    }


def _style(
    xlwt: Any,
    *,
    bold: bool = False,
    border: bool = False,
    align: str = "left",
    wrap: bool = False,
) -> Any:
    style = xlwt.XFStyle()
    font = xlwt.Font()
    font.name = "Arial"
    font.height = 100
    font.bold = bold
    style.font = font
    alignment = xlwt.Alignment()
    alignment.horz = {
        "left": xlwt.Alignment.HORZ_LEFT,
        "center": xlwt.Alignment.HORZ_CENTER,
    }.get(align, xlwt.Alignment.HORZ_LEFT)
    alignment.vert = xlwt.Alignment.VERT_CENTER
    alignment.wrap = int(wrap)
    style.alignment = alignment
    if border:
        borders = xlwt.Borders()
        borders.left = xlwt.Borders.THIN
        borders.right = xlwt.Borders.THIN
        borders.top = xlwt.Borders.THIN
        borders.bottom = xlwt.Borders.THIN
        style.borders = borders
    return style


def _write_header(sheet: Any, graph: CroquiGraph, styles: dict[str, Any]) -> None:
    header = graph.header
    sheet.write_merge(1, 3, 1, 7, "RGE", styles["note"])
    sheet.write_merge(4, 4, 1, 6, "Departamento:", styles["label"])
    sheet.write_merge(4, 4, 8, 14, header.departamento, styles["value"])
    sheet.write_merge(4, 4, 15, 21, "Município:", styles["label"])
    sheet.write_merge(4, 4, 23, 32, header.municipio, styles["value"])
    sheet.write_merge(4, 4, 33, 40, "Equipamento :", styles["label"])
    sheet.write_merge(4, 4, 41, 48, header.equipamento, styles["value"])
    sheet.write_merge(5, 5, 1, 6, "Data do Levantamento:", styles["label"])
    sheet.write_merge(5, 5, 8, 14, header.data_levantamento, styles["value"])
    sheet.write_merge(5, 5, 15, 31, "Levantamento de campo realizado por:", styles["label"])
    sheet.write_merge(5, 5, 33, 48, header.responsavel, styles["value"])


def _write_validation(sheet: Any, graph: CroquiGraph, styles: dict[str, Any]) -> None:
    sheet.write_merge(31, 31, 1, 8, "Status", styles["label"])
    sheet.write_merge(31, 31, 9, 48, graph.validation.status, styles["value"])
    messages = [*graph.validation.blockingErrors, *graph.validation.warnings]
    text = "; ".join(str(item.get("code") or item) for item in messages) or "Sem avisos."
    sheet.write_merge(32, 36, 1, 48, text, styles["warning"])


def _write_validation_report(graph: CroquiGraph, output_path: Path) -> Path:
    final_allowed = (
        graph.validation.status == "final_candidate" and not graph.validation.blockingErrors
    )
    report = {
        "contract": {
            "project_id": graph.id,
            "expected_or_selected_main_equipment": graph.header.equipamento,
            "selected_equipment_type": graph.mainEquipment.type,
            "selected_equipment_code": graph.mainEquipment.code,
            "selected_equipment_confidence": graph.mainEquipment.confidence,
            "primary_focus_code": graph.mainEquipment.code,
            "focus_validated": final_allowed,
        },
        "generated": {
            "pdf_header_equipment": graph.header.equipamento,
            "xls_header_equipment": graph.header.equipamento,
            "primary_focus_code": graph.mainEquipment.code,
            "visible_codes": [node.code for node in graph.nodes if node.code],
            "excluded_codes": [],
        },
        "validation": {
            "status": "BLOCKED" if graph.validation.blockingErrors else "DRAFT_REVIEW_REQUIRED",
            "output_status": graph.validation.status,
            "final_output_allowed": final_allowed,
            "blocking_errors": graph.validation.blockingErrors,
            "warnings": graph.validation.warnings,
        },
    }
    if final_allowed:
        report["validation"]["status"] = "PASSED"
    output_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path
